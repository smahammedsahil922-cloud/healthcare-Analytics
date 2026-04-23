"""
Healthcare Analytics Pipeline
Generates a fully-formatted multi-sheet Excel report from raw patient data.
Run: python healthcare_analysis.py
"""

import random
from datetime import date, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

# ── CONSTANTS ────────────────────────────────────────────────────────────────
DEPARTMENTS   = ["Cardiology", "Neurology", "Pediatrics", "Orthopedics", "General"]
DISEASES      = ["Diabetes", "Hypertension", "Asthma", "Back Pain", "Infection"]
INSURANCE     = ["Medicare", "Medicaid", "Private", "Self-pay"]
GENDERS       = ["Male", "Female"]

COLORS = {
    "header_dark":   "0A1628",
    "header_teal":   "0D9488",
    "header_blue":   "1A3A6B",
    "accent_cyan":   "22D3EE",
    "row_alt":       "EFF6FF",
    "row_white":     "FFFFFF",
    "positive":      "D1FAE5",
    "negative":      "FEE2E2",
    "text_dark":     "0F172A",
    "text_muted":    "64748B",
    "gold":          "F59E0B",
}

THIN = Side(style="thin", color="E2E8F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ── DATA GENERATION ──────────────────────────────────────────────────────────
def generate_data(n=500, seed=42):
    random.seed(seed)
    start = date(2023, 1, 1)
    rows = []
    for i in range(n):
        dept = random.choice(DEPARTMENTS)
        age  = random.randint(1, 90)
        cost = random.randint(500, 5000)
        rev  = random.randint(1000, 8000)
        visit_date = start + timedelta(days=random.randint(0, 499))
        rows.append({
            "Patient_ID":     1000 + i,
            "Age":            age,
            "Gender":         random.choice(GENDERS),
            "Department":     dept,
            "Disease":        random.choice(DISEASES),
            "Visit_Date":     visit_date,
            "Cost":           cost,
            "Revenue":        rev,
            "Insurance_Type": random.choice(INSURANCE),
            "Readmission":    random.choice(["Yes", "No"]),
        })
    df = pd.DataFrame(rows)
    df["Month"]  = df["Visit_Date"].apply(lambda d: d.strftime("%Y-%m"))
    df["Profit"] = df["Revenue"] - df["Cost"]
    return df


# ── STYLE HELPERS ─────────────────────────────────────────────────────────────
def hfont(bold=True, size=11, color="FFFFFF", name="Arial"):
    return Font(bold=bold, size=size, color=color, name=name)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

def currency_fmt():  return '$#,##0'
def pct_fmt():       return '0.0"%"'
def int_fmt():       return '#,##0'

def style_header_row(ws, row, cols, bg=None, fg="FFFFFF", size=11):
    bg = bg or COLORS["header_dark"]
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font      = hfont(bold=True, size=size, color=fg)
        cell.fill      = fill(bg)
        cell.alignment = center()
        cell.border    = BORDER

def style_data_row(ws, row, cols, alt=False):
    bg = COLORS["row_alt"] if alt else COLORS["row_white"]
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = fill(bg)
        cell.alignment = left()
        cell.border    = BORDER
        cell.font      = Font(name="Arial", size=10, color=COLORS["text_dark"])

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell


# ── SHEET 1 — EXECUTIVE SUMMARY ──────────────────────────────────────────────
def build_summary(wb, df):
    ws = wb.create_sheet("📊 Executive Summary")
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value     = "🏥  HEALTHCARE ANALYTICS REPORT"
    t.font      = Font(name="Arial", bold=True, size=18, color="FFFFFF")
    t.fill      = fill(COLORS["header_dark"])
    t.alignment = center()
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    sub = ws["A2"]
    sub.value     = (
        f"Generated on {date.today().strftime('%B %d, %Y')}  |  "
        f"Total Records: {len(df):,}  |  "
        f"Date Range: {df['Visit_Date'].min()}  to  {df['Visit_Date'].max()}"
    )
    sub.font      = Font(name="Arial", size=10, color=COLORS["text_muted"], italic=True)
    sub.fill      = fill("F8FAFC")
    sub.alignment = center()
    ws.row_dimensions[2].height = 22

    # KPI section header
    ws.merge_cells("A4:F4")
    kh = ws["A4"]
    kh.value     = "KEY PERFORMANCE INDICATORS"
    kh.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    kh.fill      = fill(COLORS["header_teal"])
    kh.alignment = center()
    ws.row_dimensions[4].height = 28

    kpis = [
        ("Total Patients",       len(df),                                                          int_fmt(),      "👥"),
        ("Total Revenue",        df["Revenue"].sum(),                                               currency_fmt(), "💰"),
        ("Total Cost",           df["Cost"].sum(),                                                  currency_fmt(), "💸"),
        ("Net Profit",           df["Profit"].sum(),                                                currency_fmt(), "📈"),
        ("Avg Revenue / Visit",  df["Revenue"].mean(),                                             currency_fmt(), "💳"),
        ("Avg Cost / Visit",     df["Cost"].mean(),                                                currency_fmt(), "🏷"),
        ("Profit Margin",        df["Profit"].sum() / df["Revenue"].sum() * 100,                   '0.0"%"',       "📊"),
        ("Readmission Rate",     (df["Readmission"] == "Yes").mean() * 100,                        '0.0"%"',       "🔁"),
        ("Avg Patient Age",      df["Age"].mean(),                                                  "0.0",          "🧑"),
        ("Departments",          df["Department"].nunique(),                                        "0",            "🏨"),
    ]

    headers = ["Icon", "KPI Metric", "Value", "Status", "Benchmark", "Notes"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font      = hfont(size=10, color="FFFFFF")
        cell.fill      = fill(COLORS["header_blue"])
        cell.alignment = center()
        cell.border    = BORDER

    for i, (label, val, fmt, icon) in enumerate(kpis):
        r   = 6 + i
        alt = i % 2 == 1

        ws.cell(r, 1, icon).alignment   = center()
        ws.cell(r, 2, label)
        ws.cell(r, 3, round(val, 2) if isinstance(val, float) else val)
        ws.cell(r, 3).number_format     = fmt
        ws.cell(r, 4, "✅ On Target")
        ws.cell(r, 5, "–")
        ws.cell(r, 6, "Auto-calculated")

        style_data_row(ws, r, 6, alt=alt)
        ws.cell(r, 1).alignment = center()
        ws.cell(r, 3).alignment = center()
        ws.cell(r, 4).alignment = center()
        ws.row_dimensions[r].height = 22

    set_col_widths(ws, [8, 28, 18, 18, 18, 24])
    freeze(ws, "A5")


# ── SHEET 2 — DEPARTMENT ANALYSIS ────────────────────────────────────────────
def build_department(wb, df):
    ws = wb.create_sheet("🏨 Department Analysis")
    ws.sheet_view.showGridLines = False

    dept = (
        df.groupby("Department").agg(
            Patients=("Patient_ID", "count"),
            Avg_Age=("Age", "mean"),
            Total_Revenue=("Revenue", "sum"),
            Total_Cost=("Cost", "sum"),
            Avg_Revenue=("Revenue", "mean"),
            Avg_Cost=("Cost", "mean"),
            Net_Profit=("Profit", "sum"),
            Readmissions=("Readmission", lambda x: (x == "Yes").sum()),
        )
        .reset_index()
    )
    dept["Readmission_Rate"] = dept["Readmissions"] / dept["Patients"] * 100

    # Title
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value     = "DEPARTMENT-WISE PERFORMANCE"
    t.font      = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill      = fill(COLORS["header_dark"])
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    headers = [
        "Department", "Patients", "Avg Age", "Total Revenue ($)",
        "Total Cost ($)", "Avg Revenue ($)", "Avg Cost ($)",
        "Net Profit ($)", "Readmission Rate (%)"
    ]
    style_header_row(ws, 2, len(headers), bg=COLORS["header_teal"])
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)

    cols_fmt = [None, int_fmt(), "0.0", currency_fmt(), currency_fmt(),
                currency_fmt(), currency_fmt(), currency_fmt(), pct_fmt()]

    for i, row in dept.iterrows():
        r   = 3 + i
        vals = [
            row["Department"], row["Patients"], round(row["Avg_Age"], 1),
            row["Total_Revenue"], row["Total_Cost"], row["Avg_Revenue"],
            row["Avg_Cost"], row["Net_Profit"], round(row["Readmission_Rate"], 1)
        ]
        style_data_row(ws, r, len(vals), alt=i % 2 == 1)
        for c, (v, fmt) in enumerate(zip(vals, cols_fmt), 1):
            cell = ws.cell(r, c, round(v, 2) if isinstance(v, float) else v)
            if fmt:
                cell.number_format = fmt
            cell.alignment = left() if c == 1 else center()
        ws.row_dimensions[r].height = 22

    # Totals row
    tr = 3 + len(dept)
    style_header_row(ws, tr, len(headers), bg=COLORS["header_blue"])
    ws.cell(tr, 1, "TOTAL").font = hfont(size=10)
    for c, col in [(2, "Patients"), (4, "Total_Revenue"), (5, "Total_Cost"), (8, "Net_Profit")]:
        ws.cell(tr, c, dept[col].sum()).number_format = currency_fmt() if c != 2 else int_fmt()
        ws.cell(tr, c).font = hfont(size=10)
        ws.cell(tr, c).alignment = center()

    # Color scale on profit column
    profit_range = f"H3:H{tr-1}"
    ws.conditional_formatting.add(
        profit_range,
        ColorScaleRule(
            start_type="min", start_color="FEE2E2",
            mid_type="percentile", mid_value=50, mid_color="FFFFFF",
            end_type="max", end_color="D1FAE5"
        )
    )

    set_col_widths(ws, [18, 12, 10, 20, 18, 18, 16, 18, 22])
    freeze(ws, "A3")


# ── SHEET 3 — INSURANCE ANALYSIS ─────────────────────────────────────────────
def build_insurance(wb, df):
    ws = wb.create_sheet("🛡 Insurance Analysis")
    ws.sheet_view.showGridLines = False

    ins = (
        df.groupby("Insurance_Type").agg(
            Patients=("Patient_ID", "count"),
            Total_Revenue=("Revenue", "sum"),
            Total_Cost=("Cost", "sum"),
            Avg_Revenue=("Revenue", "mean"),
            Net_Profit=("Profit", "sum"),
            Readmissions=("Readmission", lambda x: (x == "Yes").sum()),
        ).reset_index()
    )
    ins["Readmission_Rate"] = ins["Readmissions"] / ins["Patients"] * 100

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "INSURANCE TYPE ANALYSIS"
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = fill(COLORS["header_dark"])
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    headers = ["Insurance Type", "Patients", "Total Revenue ($)", "Total Cost ($)",
               "Avg Revenue ($)", "Net Profit ($)", "Readmission Rate (%)"]
    style_header_row(ws, 2, len(headers), bg=COLORS["header_teal"])
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)

    fmts = [None, int_fmt(), currency_fmt(), currency_fmt(), currency_fmt(), currency_fmt(), pct_fmt()]
    for i, row in ins.iterrows():
        r = 3 + i
        vals = [row["Insurance_Type"], row["Patients"], row["Total_Revenue"],
                row["Total_Cost"], row["Avg_Revenue"], row["Net_Profit"],
                round(row["Readmission_Rate"], 1)]
        style_data_row(ws, r, len(vals), alt=i % 2 == 1)
        for c, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r, c, round(v, 2) if isinstance(v, float) else v)
            if fmt: cell.number_format = fmt
            cell.alignment = left() if c == 1 else center()
        ws.row_dimensions[r].height = 22

    set_col_widths(ws, [18, 12, 20, 18, 18, 18, 22])
    freeze(ws, "A3")


# ── SHEET 4 — DISEASE BREAKDOWN ──────────────────────────────────────────────
def build_disease(wb, df):
    ws = wb.create_sheet("🦠 Disease Analysis")
    ws.sheet_view.showGridLines = False

    dis = (
        df.groupby("Disease").agg(
            Patients=("Patient_ID", "count"),
            Avg_Age=("Age", "mean"),
            Avg_Cost=("Cost", "mean"),
            Avg_Revenue=("Revenue", "mean"),
            Readmissions=("Readmission", lambda x: (x == "Yes").sum()),
        ).reset_index()
    )
    dis["Readmission_Rate"] = dis["Readmissions"] / dis["Patients"] * 100
    dis = dis.sort_values("Patients", ascending=False).reset_index(drop=True)

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "DISEASE-WISE ANALYSIS"
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = fill(COLORS["header_dark"])
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    headers = ["Disease", "Patients", "Avg Age", "Avg Cost ($)", "Avg Revenue ($)", "Readmission Rate (%)"]
    style_header_row(ws, 2, len(headers), bg=COLORS["header_teal"])
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)

    fmts = [None, int_fmt(), "0.0", currency_fmt(), currency_fmt(), pct_fmt()]
    for i, row in dis.iterrows():
        r = 3 + i
        vals = [row["Disease"], row["Patients"], round(row["Avg_Age"], 1),
                round(row["Avg_Cost"], 0), round(row["Avg_Revenue"], 0),
                round(row["Readmission_Rate"], 1)]
        style_data_row(ws, r, len(vals), alt=i % 2 == 1)
        for c, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r, c, v)
            if fmt: cell.number_format = fmt
            cell.alignment = left() if c == 1 else center()
        ws.row_dimensions[r].height = 22

    set_col_widths(ws, [18, 12, 10, 18, 18, 22])
    freeze(ws, "A3")


# ── SHEET 5 — MONTHLY TRENDS ─────────────────────────────────────────────────
def build_monthly(wb, df):
    ws = wb.create_sheet("📈 Monthly Trends")
    ws.sheet_view.showGridLines = False

    monthly = (
        df.groupby("Month").agg(
            Visits=("Patient_ID", "count"),
            Total_Revenue=("Revenue", "sum"),
            Total_Cost=("Cost", "sum"),
            Net_Profit=("Profit", "sum"),
        ).reset_index().sort_values("Month")
    )

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "MONTHLY REVENUE & COST TRENDS"
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = fill(COLORS["header_dark"])
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    headers = ["Month", "Visits", "Total Revenue ($)", "Total Cost ($)", "Net Profit ($)"]
    style_header_row(ws, 2, 5, bg=COLORS["header_teal"])
    for c, h in enumerate(headers, 1):
        ws.cell(2, c, h)

    fmts = [None, int_fmt(), currency_fmt(), currency_fmt(), currency_fmt()]
    for i, row in monthly.iterrows():
        r = 3 + i
        vals = [row["Month"], row["Visits"], row["Total_Revenue"], row["Total_Cost"], row["Net_Profit"]]
        style_data_row(ws, r, 5, alt=i % 2 == 1)
        for c, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r, c, v)
            if fmt: cell.number_format = fmt
            cell.alignment = center()
        ws.row_dimensions[r].height = 22

        # Colour profit cell
        profit_cell = ws.cell(r, 5)
        if row["Net_Profit"] > 0:
            profit_cell.fill = fill(COLORS["positive"])
        else:
            profit_cell.fill = fill(COLORS["negative"])

    set_col_widths(ws, [14, 10, 22, 20, 20])
    freeze(ws, "A3")


# ── SHEET 6 — DEMOGRAPHICS ───────────────────────────────────────────────────
def build_demographics(wb, df):
    ws = wb.create_sheet("👥 Demographics")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "DEMOGRAPHIC ANALYSIS"
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = fill(COLORS["header_dark"])
    t.alignment = center()
    ws.row_dimensions[1].height = 36

    # Gender
    gender = (
        df.groupby("Gender").agg(
            Patients=("Patient_ID", "count"),
            Avg_Age=("Age", "mean"),
            Avg_Cost=("Cost", "mean"),
            Avg_Revenue=("Revenue", "mean"),
            Readmissions=("Readmission", lambda x: (x == "Yes").sum()),
        ).reset_index()
    )
    gender["Readmission_Rate"] = gender["Readmissions"] / gender["Patients"] * 100

    gh = ["Gender", "Patients", "Avg Age", "Avg Cost ($)", "Avg Revenue ($)", "Readmission Rate (%)"]
    style_header_row(ws, 2, 6, bg=COLORS["header_blue"])
    for c, h in enumerate(gh, 1):
        ws.cell(2, c, h)

    for i, row in gender.iterrows():
        r = 3 + i
        vals = [row["Gender"], row["Patients"], round(row["Avg_Age"], 1),
                round(row["Avg_Cost"], 0), round(row["Avg_Revenue"], 0),
                round(row["Readmission_Rate"], 1)]
        style_data_row(ws, r, 6, alt=i % 2 == 1)
        fmts = [None, int_fmt(), "0.0", currency_fmt(), currency_fmt(), pct_fmt()]
        for c, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r, c, v)
            if fmt: cell.number_format = fmt
            cell.alignment = left() if c == 1 else center()
        ws.row_dimensions[r].height = 22

    # Age groups
    df["Age_Group"] = pd.cut(
        df["Age"], bins=[0, 17, 35, 55, 75, 200],
        labels=["0–17", "18–35", "36–55", "56–75", "76+"]
    )
    age_grp = (
        df.groupby("Age_Group", observed=True).agg(
            Patients=("Patient_ID", "count"),
            Avg_Revenue=("Revenue", "mean"),
            Avg_Cost=("Cost", "mean"),
        ).reset_index()
    )

    ws.merge_cells("A6:F6")
    ag_title = ws["A6"]
    ag_title.value = "AGE GROUP BREAKDOWN"
    ag_title.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    ag_title.fill = fill(COLORS["header_teal"])
    ag_title.alignment = center()
    ws.row_dimensions[6].height = 26

    agh = ["Age Group", "Patients", "Avg Revenue ($)", "Avg Cost ($)"]
    style_header_row(ws, 7, 4, bg=COLORS["header_blue"])
    for c, h in enumerate(agh, 1):
        ws.cell(7, c, h)

    for i, row in age_grp.iterrows():
        r = 8 + i
        vals = [str(row["Age_Group"]), row["Patients"], round(row["Avg_Revenue"], 0), round(row["Avg_Cost"], 0)]
        style_data_row(ws, r, 4, alt=i % 2 == 1)
        fmts = [None, int_fmt(), currency_fmt(), currency_fmt()]
        for c, (v, fmt) in enumerate(zip(vals, fmts), 1):
            cell = ws.cell(r, c, v)
            if fmt: cell.number_format = fmt
            cell.alignment = left() if c == 1 else center()
        ws.row_dimensions[r].height = 22

    set_col_widths(ws, [14, 12, 18, 16, 18, 22])
    freeze(ws, "A3")


# ── SHEET 7 — RAW DATA ───────────────────────────────────────────────────────
def build_raw(wb, df):
    ws = wb.create_sheet("📋 Raw Data")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value = "PATIENT VISIT DATA"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = fill(COLORS["header_dark"])
    t.alignment = center()
    ws.row_dimensions[1].height = 32

    export_cols = ["Patient_ID","Age","Gender","Department","Disease",
                   "Visit_Date","Cost","Revenue","Insurance_Type","Readmission","Month","Profit"]
    style_header_row(ws, 2, len(export_cols), bg=COLORS["header_teal"])
    for c, h in enumerate(export_cols, 1):
        ws.cell(2, c, h)

    date_fmt = "YYYY-MM-DD"
    for i, (_, row) in enumerate(df[export_cols].iterrows()):
        r = 3 + i
        alt = i % 2 == 1
        style_data_row(ws, r, len(export_cols), alt=alt)
        for c, col in enumerate(export_cols, 1):
            val = row[col]
            cell = ws.cell(r, c, val)
            cell.alignment = center()
            if col in ("Cost", "Revenue", "Profit"):
                cell.number_format = currency_fmt()
            elif col == "Visit_Date":
                cell.number_format = date_fmt

    set_col_widths(ws, [12, 8, 10, 16, 16, 14, 12, 12, 16, 14, 10, 12])
    freeze(ws, "A3")


# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    print("🔄 Generating patient data...")
    df = generate_data(500)

    print("📊 Building Excel report...")
    wb = Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    build_summary(wb, df)
    build_department(wb, df)
    build_insurance(wb, df)
    build_disease(wb, df)
    build_monthly(wb, df)
    build_demographics(wb, df)
    build_raw(wb, df)

    out = "Healthcare_Analysis_Report.xlsx"
    wb.save(out)
    print(f"✅ Report saved → {out}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
